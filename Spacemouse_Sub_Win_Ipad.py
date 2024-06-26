#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Jun  2 02:38:10 2024
@author: kathanpatel sub_Ipad.py
"""
import socket
import json
import time
import keyboard  # Import the keyboard module
import threading  # Import threading module
from collections import deque  # Import deque for efficient storage of joint angles
from openpyxl import Workbook, load_workbook

waypoints = {}  # Initialize waypoints dictionary globally
correcting_orientation = False
first_home = False


def connectETController(ip, port=8055):
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        sock.connect((ip, port))
        return True, sock
    except Exception as e:
        sock.close()
        return False, None

def disconnectETController(sock):
    if sock:
        sock.close()

def sendCMD(sock, cmd, params=None, id=1):
    if not params:
        params = []
    else:
        params = json.dumps(params)
    sendStr = '{{"method":"{0}","params":{1},"jsonrpc":"2.0","id":{2}}}\n'.format(cmd, params, id)
    try:
        sock.sendall(sendStr.encode('utf-8'))
        ret = sock.recv(1024)
        jdata = json.loads(ret.decode('utf-8'))
        if 'result' in jdata:
            return True, json.loads(jdata['result']), jdata['id']
        elif 'error' in jdata:
            return False, jdata['error'], jdata['id']
        else:
            return False, None, None
    except Exception as e:
        return False, None, None

def set_v(data, robot_speed, omega):
    global current_pose
    
    for i in range(6):
        if data[i] == 1:
            current_pose[i] = robot_speed
        elif data[i] == -1:
            current_pose[i] = -robot_speed
        else:
            current_pose[i] = 0        
    
    return current_pose

def calculate_inverse_kinematics(robot_ip, target_pose):
    P000 = [0, -90, -90, 90, -90, 0]

    conSuc, sock = connectETController(robot_ip)

    if conSuc:
        try:
            suc, result, id = sendCMD(sock, "inverseKinematic", {"targetPose": target_pose, "referencePos": P000})
            if suc:
                return result
            else:
                print("Inverse kinematics failed for this pose:", result)
                return None

        except Exception as e:
            print("Error:", e)
        finally:
            disconnectETController(sock)
    else:
        print("Connection to the robot failed.")
        return None

def Orientation_correct(robot_ip):
    conSuc, sock = connectETController(robot_ip)

    if not conSuc:
        return None

    suc, result, id = sendCMD(sock, "set_servo_status", {"status": 1})
    suc, Current_tcp, id = sendCMD(sock, 'get_tcp_pose', {'coordinate_num': 0, 'tool_num': 0})
    
    points = [Current_tcp[0], Current_tcp[1], Current_tcp[2], -1.57,0,0]
    
    #print("This is the desired point in coordinate system" , points)
    
    angle_point = calculate_inverse_kinematics(robot_ip, points)
    
    angle_point[5] = angle_point[5] - 90

    print("Moving to point linearly:", angle_point)

    if angle_point is None:
        print("Error: Target position for linear motion is invalid.")
        return

    suc, result, id = sendCMD(sock, "moveByLine", {
        "targetPos": angle_point,
        "speed_type": 0,
        "speed": 200,
        "cond_type": 0,
        "cond_num": 7,
        "cond_value": 1})

    if not suc:
        print("Error in moveByLine:", result)
        return


    while True:
        suc, result, id = sendCMD(sock, "getRobotState")
        if result == 0:
            break


def wrist3_calibate(robot_ip):
    conSuc, sock = connectETController(robot_ip)

    if not conSuc:
        return None

    suc, result, id = sendCMD(sock, "set_servo_status", {"status": 1})
    suc, Joint_angles, id = sendCMD(sock, "get_joint_pos")
    Joint_angles[5] = 0

    print("Moving to point linearly:", Joint_angles)

    if Joint_angles is None:
        print("Error: Target position for linear motion is invalid.")
        return

    suc, result, id = sendCMD(sock, "moveByJoint", {
        "targetPos": Joint_angles,
        "speed": 30,
        "acc":10,
        "dec":10,
        "cond_type": 0,
        "cond_num": 7,
        "cond_value": 1})

    if not suc:
        print("Error in moveByLine:", result)
        return


    while True:
        suc, result, id = sendCMD(sock, "getRobotState")
        if result == 0:
            break

def Parking(robot_ip):
    conSuc, sock = connectETController(robot_ip)

    if not conSuc:
        return None

    suc, result, id = sendCMD(sock, "set_servo_status", {"status": 1})
    
    
    points = [217, 648, 71, -2.14,0,0]
    
    #print("This is the desired point in coordinate system" , points)
    
    angle_point = calculate_inverse_kinematics(robot_ip, points)
    
    angle_point[5] = angle_point[5] - 90

    print("Moving to point linearly:", angle_point)

    if angle_point is None:
        print("Error: Target position for linear motion is invalid.")
        return

    suc, result, id = sendCMD(sock, "moveByLine", {
        "targetPos": angle_point,
        "speed_type": 0,
        "speed": 100,
        "cond_type": 0,
        "cond_num": 7,
        "cond_value": 1})

    if not suc:
        print("Error in moveByLine:", result)
        return


    while True:
        suc, result, id = sendCMD(sock, "getRobotState")
        if result == 0:
            break
def homing(robot_ip):
    global first_home
    global homing_coord
    
    
    conSuc, sock = connectETController(robot_ip)

    if not conSuc:
        return None

    suc, result, id = sendCMD(sock, "set_servo_status", {"status": 1})
    suc, Current_tcp, id = sendCMD(sock, 'get_tcp_pose', {'coordinate_num': 0, 'tool_num': 0})
    
    if not first_home:
        suc, homing_coord, id = sendCMD(sock, 'get_tcp_pose', {'coordinate_num': 0, 'tool_num': 0})
        first_home = True
        print("Saved new home location!!!")
     
    #print("This is the desired point in coordinate system", homing_coord)
    
    angle_point = calculate_inverse_kinematics(robot_ip, homing_coord)

    print("Moving to point linearly:", angle_point)

    if angle_point is None:
        print("Error: Target position for linear motion is invalid.")
        return

    suc, result, id = sendCMD(sock, "moveByLine", {
        "targetPos": angle_point,
        "speed_type": 0,
        "speed": 100,
        "cond_type": 0,
        "cond_num": 7,
        "cond_value": 1})

    if not suc:
        print("Error in moveByLine:", result)
        return
    
    

    while True:
        suc, result, id = sendCMD(sock, "getRobotState")
        if result == 0:
            break
        
def get_waypoint_from_excel(excel_file, waypoint_key):
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
    except FileNotFoundError:
        print(f"Excel file '{excel_file}' not found.")
        return None
    
    for row in ws.iter_rows(min_row=2, max_col=7):  # Adjust the range to include the necessary columns
        if row[0].value == waypoint_key:
            coordinates = [
                row[1].value,  # X
                row[2].value,  # Y
                row[3].value,  # Z
                row[4].value,  # RX
                row[5].value,  # RY
                row[6].value   # RZ
            ]
            return coordinates
    
    return None


def save_waypoint_to_excel(excel_file, waypoint_key, coordinates):
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(['Waypoint Key', 'X', 'Y', 'Z', 'RX', 'RY', 'RZ'])

    found = False
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if row[0] == waypoint_key:
            ws.cell(row=row[0], column=2).value = coordinates[0]  # X
            ws.cell(row=row[0], column=3).value = coordinates[1]  # Y
            ws.cell(row=row[0], column=4).value = coordinates[2]  # Z
            ws.cell(row=row[0], column=5).value = coordinates[3]  # RX
            ws.cell(row=row[0], column=6).value = coordinates[4]  # RY
            ws.cell(row=row[0], column=7).value = coordinates[5]  # RZ
            found = True
            break
    
    if not found:
        ws.append([waypoint_key] + coordinates)
    
    wb.save(excel_file)
    print(f"Waypoint {waypoint_key} saved to {excel_file}")

def initialize_excel_file(excel_file):
    """Initializes the Excel file by creating it with the necessary headers."""
    wb = Workbook()
    ws = wb.active
    ws.append(['Waypoint Key', 'X', 'Y', 'Z', 'RX', 'RY', 'RZ'])  # Add headers
    wb.save(excel_file)
    print(f"Excel file '{excel_file}' initialized.")
    
def delete_waypoint_from_excel(waypoint_key):
    excel_file = r"C:\Users\X'ian\Downloads\Windows_Spacemouse_control_stack-main\waypoints.xlsx"
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
    except FileNotFoundError:
        print(f"Excel file '{excel_file}' not found.")
        return
    
    found = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == waypoint_key:
            ws.delete_rows(row)
            found = True
            break
    
    if not found:
        print(f"Waypoint {waypoint_key} not found in {excel_file}.")
    else:
        wb.save(excel_file)
        print(f"Waypoint {waypoint_key} deleted from {excel_file}")

    
def waypoint(robot_ip, waypoint_key):
    global correcting_orientation
    conSuc, sock = connectETController(robot_ip)

    if not conSuc:
        return None

    suc, result, id = sendCMD(sock, "set_servo_status", {"status": 1})
    suc, Current_tcp, id = sendCMD(sock, 'get_tcp_pose', {'coordinate_num': 0, 'tool_num': 0})
    
    excel_file = r"C:\Users\X'ian\Downloads\Windows_Spacemouse_control_stack-main\waypoints.xlsx"
    coordinates = get_waypoint_from_excel(excel_file, waypoint_key)
    
    if coordinates is None:
        # Waypoint not found in Excel, add it
        coordinates = Current_tcp
        save_waypoint_to_excel(excel_file, waypoint_key, coordinates)
        print(f"Saved new waypoint {waypoint_key} location!!!")
    else:
        print(f"Found waypoint {waypoint_key} in Excel. Moving to coordinates.")

    # Now coordinates should have the correct waypoint coordinates
    angle_point = calculate_inverse_kinematics(robot_ip, coordinates)

    print("Moving to point linearly:", angle_point)

    if angle_point is None:
        print("Error: Target position for linear motion is invalid.")
        return

    suc, result, id = sendCMD(sock, "moveByLine", {
        "targetPos": angle_point,
        "speed_type": 0,
        "speed": 70,
        "cond_type": 0,
        "cond_num": 7,
        "cond_value": 1})

    if not suc:
        print("Error in moveByLine:", result)
        return

    while True:
        suc, result, id = sendCMD(sock, "getRobotState")
        if result == 0:
            break

def Waypoint_trcking(robot_ip,joint_angles):
    conSuc, sock = connectETController(robot_ip)

    if not conSuc:
        return None

    suc, result, id = sendCMD(sock, "set_servo_status", {"status": 1})
    print("Moving to point linearly:", joint_angles)
    
    if joint_angles is None:
        print("Error: Target position for linear motion is invalid.")
        return

    suc, result, id = sendCMD(sock, "moveByLine", {
        "targetPos": joint_angles,
        "speed_type": 0,
        "speed": 100,
        "cond_type": 0,
        "cond_num": 7,
        "cond_value": 1})

    if not suc:
        print("Error in moveByLine:", result)
        return

    while True:
        suc, result, id = sendCMD(sock, "getRobotState")
        if result == 0:
            break

def keypress_handler(robot_ip, joint_angles_deque):
    global correcting_orientation
    global first_home
    
    while True:
        if keyboard.is_pressed('1') and not correcting_orientation:
            correcting_orientation = True
            waypoint(robot_ip, 1)
            print("Reached waypoint 1")
            correcting_orientation = False
        elif keyboard.is_pressed('2'):
            correcting_orientation = True
            waypoint(robot_ip, 2)
            print("Reached waypoint 2")
            correcting_orientation = False
        elif keyboard.is_pressed('3'):
            correcting_orientation = True
            waypoint(robot_ip, 3)
            print("Reached waypoint 3")
            correcting_orientation = False
        elif keyboard.is_pressed('4'):
            correcting_orientation = True
            waypoint(robot_ip, 4)
            print("Reached waypoint 4")
            correcting_orientation = False
        elif keyboard.is_pressed('5'):
            correcting_orientation = True
            waypoint(robot_ip, 5)
            print("Reached waypoint 5")
            correcting_orientation = False
        elif keyboard.is_pressed('6'):
            correcting_orientation = True
            waypoint(robot_ip, 6)
            print("Reached waypoint 6")
            correcting_orientation = False
        elif keyboard.is_pressed('7'):
            correcting_orientation = True
            waypoint(robot_ip, 7)
            print("Reached waypoint 7")
            correcting_orientation = False
        elif keyboard.is_pressed('8'):
            correcting_orientation = True
            waypoint(robot_ip, 8)
            print("Reached waypoint 8")
            correcting_orientation = False
        elif keyboard.is_pressed('9'):
            correcting_orientation = True
            waypoint(robot_ip, 9)
            print("Reached waypoint 9")
            correcting_orientation = False
        if keyboard.is_pressed('O') and not correcting_orientation:
            print("Received 'o' key press")  # Press O to recalibrate joints to an right orientation, Watch out for any possible collision
            correcting_orientation = True
            Orientation_correct(robot_ip)
            print("Orientation corrected!")
            correcting_orientation = False
        if keyboard.is_pressed('C') and not correcting_orientation:
            print("Received 'c' key press")  # Press C to recalibrate joint 6 to 0 degree , Watch out for any possible collision
            correcting_orientation = True   
            wrist3_calibate(robot_ip)
            print("Orientation corrected!")
            correcting_orientation = False
        if keyboard.is_pressed('P') and not correcting_orientation:
            print("Received 'p' key press")  # Press P to make robot park to safe place
            correcting_orientation = True
            Parking(robot_ip)
            print("Reached Parking pose!")
            correcting_orientation = False
        if keyboard.is_pressed('H') and not correcting_orientation:
            print("Received 'h' key press")  # Press H to take robot to user define Patient home
            correcting_orientation = True
            homing(robot_ip)
            print("Reached Home!")
            correcting_orientation = False
        if keyboard.is_pressed('T') :    # Press 'T' to undo robot motion to 5 seconds before.
            correcting_orientation = True
            if len(joint_angles_deque) >= 315:
                joint_angles_315_cycles_ago, cycle_time = joint_angles_deque[-315]
                Waypoint_trcking(robot_ip,joint_angles_315_cycles_ago)
                print("Joint angles 315 cycles ago:", joint_angles_315_cycles_ago)
                print(f"Time taken for one cycle 315 cycles ago: {cycle_time:.6f} seconds")
            else:
                print("Not enough data to provide joint angles from 315 cycles ago.")
            correcting_orientation = False
        if keyboard.is_pressed('r'):
            waypoint_key = input("Enter the waypoint key to delete (1-9), press 'a' to delete all waypoints, or press 'h' to delete patient home: ")
            if waypoint_key.isdigit() and int(waypoint_key) in range(1, 10):
                delete_waypoint_from_excel(int(waypoint_key))
            elif waypoint_key.lower() == 'a':
                #print("You have pressed the 'a' key after pressing 'r'.")
                excel_file = r"C:\Users\X'ian\Downloads\Windows_Spacemouse_control_stack-main\waypoints.xlsx"  # Replace with your user directory
                initialize_excel_file(excel_file)
                print("All waypoints deleted !")
                # Add any additional actions you want to perform when 'a' is pressed
            elif waypoint_key.lower() == 'h':
                first_home = False
                print("Paitent Home waypoints deleted !")
                
            else:
                print("Invalid input. Please enter a number from 1 to 9 or press 'a' or 'h'")

        time.sleep(0.1)  # Small delay to reduce CPU usage

def main():
    global robot_speed
    global omega
    global current_pose
    global correcting_orientation
    global first_home
    global waypoints
    
    waypoints = {}
    first_home = False
    
    
    robot_speed = 10
    omega = 10
    current_pose = [0] * 8  # Initialize current_pose array
    final_matrix = [0] * 6
    mode = 0  # Initialize mode
    correcting_orientation = False
    joint_angles_deque = deque(maxlen=315)  # Initialize deque with max length of 315

    robot_ip = '192.168.1.200'
    conSuc, robot_sock = connectETController(robot_ip)
    if not conSuc:
        print('Failed to connect to the robot.')
        return

    sendCMD(robot_sock, 'set_servo_status', {'status': 1})
    # Set the loop mode to single loop
    suc, result, id = sendCMD(robot_sock, 'setCycleMode', {'cycle_mode': 2})
    suc, result, id = sendCMD(robot_sock, 'setCurrentCoord', {'coord_mode': 1})

    if conSuc:
        suc, result_pose, id = sendCMD(robot_sock, 'get_tcp_pose', {'coordinate_num': 0, 'tool_num': 0})
        print(result_pose)
        suc, result, id = sendCMD(robot_sock, 'setSysVarV', {'addr': 1, 'pose': current_pose})

    client_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    host = '127.0.0.1'
    port = 12345
    client_socket.connect((host, port))

    last_safety_check = time.time()

    # Start the keypress handler thread
    keypress_thread = threading.Thread(target=keypress_handler, args=(robot_ip, joint_angles_deque))
    keypress_thread.daemon = True  # Daemonize thread to ensure it exits when the main program exits
    keypress_thread.start()

    try:
        while True:
            cycle_start_time = time.time()

            if correcting_orientation:
                time.sleep(0.1)  # Give time for orientation correction
                continue

            data = client_socket.recv(1024)
            if not data:
                decoded_data = [0] * 8
            else:
                try:
                    decoded_data = json.loads(data.decode('utf-8'))
                except json.decoder.JSONDecodeError as e:
                    print("Error decoding JSON:", e)
                    print("Push or Pull space mouse a bit harder!!!")
                    decoded_data = [0] * 8
                    continue  # Skip processing if decoding fails

            for i in range(6):
                if i < 3:
                    if decoded_data[i] > 100:
                        decoded_data[i] = 1
                    elif decoded_data[i] < -100:
                        decoded_data[i] = -1
                    else:
                        decoded_data[i] = 0
                else:
                    if decoded_data[i] > 130:
                        decoded_data[i] = 1
                    elif decoded_data[i] < -130:
                        decoded_data[i] = -1
                    else:
                        decoded_data[i] = 0
            #Safteyy = 1

            # Perform safety check less frequently (e.g., every 0.5 seconds)
            if time.time() - last_safety_check > 0.1:
                suc, Safteyy, id = sendCMD(robot_sock, 'getVirtualOutput', {'addr': 440})
                last_safety_check = time.time()
       
            suc, Saftey_joint, id = sendCMD(robot_sock, 'getVirtualOutput', {'addr': 528})
            suc, Joint_angles, id = sendCMD(robot_sock, "get_joint_pos")

            kkp = 1 

            if (Saftey_joint == 1 or 
                Joint_angles[5] >= 340 or Joint_angles[5] <= -340 or 
                Joint_angles[3] >= 340 or Joint_angles[3] <= -340 or 
                Joint_angles[4] >= 340 or Joint_angles[4] <= -340):

                suc, result, id = sendCMD(robot_sock, "set_servo_status", {"status": 0})
                kkp = 0
            if Safteyy == 0:
                print("Robot out of the safe zone!!! Please take it back to safe zone")
            #print("Safteyy = ",Safteyy)
            if decoded_data != [0] * 8 and Safteyy != 0 and Saftey_joint == 0 and kkp == 1:
                if decoded_data[6] == 1 and decoded_data[7] == 1 and not correcting_orientation:
                    mode = 1
                    correcting_orientation = True  # Set flag before calling orientation correction
                    print(f"Switched to mode {mode}")
                    Orientation_correct(robot_ip)
                    correcting_orientation = False  # Reset flag after correction
                    mode = 0  # Reset mode after correction

                if decoded_data[6] == 1 and robot_speed > 5 and omega > 2:
                    robot_speed -= 5
                    omega -= 2
                    print('Speed Decreased to:', robot_speed)
                elif decoded_data[7] == 1 and robot_speed < 200 and omega < 150:
                    robot_speed += 5
                    omega += 2
                    print('Speed increased to:', robot_speed)
                else:
                    temp = set_v(decoded_data, robot_speed, omega)
                    final_matrix = temp[:6]
                    final_matrix[0] = final_matrix[0]
                    final_matrix[1] = -final_matrix[1]
                    final_matrix[2] = -final_matrix[2]
                    final_matrix[3] = final_matrix[3]
                    final_matrix[4] = -final_matrix[4]
                    final_matrix[5] = -final_matrix[5]
                    

                    if len(decoded_data) == 8 and not correcting_orientation:
                        print(f'Received: {final_matrix}')
                        suc, result, id = sendCMD(robot_sock, 'moveBySpeedl', {'v': final_matrix, 'acc': 40, 'arot': 10, 't': 0.07})
                        print(suc, result, id)
                        #print("Joint Angles = ", Joint_angles)
                        joint_angles_deque.append((list(Joint_angles), time.time() - cycle_start_time))
            else:
                suc, result, id = sendCMD(robot_sock, "stopl", {"acc": 690})
                decoded_data = [0] * 8

    except KeyboardInterrupt:
        print('Client stopped.')
    finally:
        client_socket.close()
        disconnectETController(robot_sock)

if __name__ == '__main__':
     main()
